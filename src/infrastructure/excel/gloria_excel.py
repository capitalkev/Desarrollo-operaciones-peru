import io
import logging
from datetime import datetime
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class GloriaExcelService:
    def generar_excel(
        self, facturas: list[dict[str, Any]]
    ) -> tuple[str | None, bytes | None]:
        """
        Genera un archivo Excel parametrizado para Gloria.
        Asume que la lista 'facturas' que recibe ya fue validada previamente.
        """
        if not facturas:
            return None, None

        data_rows = []
        fecha_envio = datetime.now().strftime("%d.%m.%Y")

        for invoice in facturas:
            document_id = str(invoice.get("document_id", ""))
            if "-" in document_id:
                parts = document_id.split("-")
                formatted_document_id = f"{parts[0]}-{parts[1].zfill(8)}"
            else:
                formatted_document_id = document_id

            # Parseo seguro de fechas con pandas (evita errores si el campo viene vacío)
            issue_date = pd.to_datetime(invoice.get("issue_date"), errors="coerce")
            due_date = pd.to_datetime(invoice.get("due_date"), errors="coerce")

            row = {
                "FACTOR": "20603596294",
                "FECHA DE ENVIO": fecha_envio,
                "RUC PROVEEDOR": invoice.get("client_ruc"),
                "PROVEEDOR": invoice.get("client_name"),
                "RUC CLIENTE": invoice.get("debtor_ruc"),
                "CLIENTE": (invoice.get("debtor_name") or "")
                .replace("-", " ")
                .replace(".", ""),
                "FECHA DE EMISION": (
                    issue_date.strftime("%d.%m.%Y") if pd.notna(issue_date) else ""
                ),
                "NUM FACTURA": formatted_document_id,
                "IMPORTE NETO PAGAR": invoice.get("net_amount", 0.0),
                "MONEDA": invoice.get("currency"),
                "FECHA DE VENCIMIENTO": (
                    due_date.strftime("%d.%m.%Y") if pd.notna(due_date) else ""
                ),
            }
            data_rows.append(row)

        df = pd.DataFrame(data_rows)
        output_buffer = io.BytesIO()

        with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Facturas", index=False)
            worksheet = writer.sheets["Facturas"]

            header_font = Font(bold=True, color="000000")
            thin_border_side = Side(border_style="thin", color="000000")
            cell_border = Border(
                left=thin_border_side,
                right=thin_border_side,
                top=thin_border_side,
                bottom=thin_border_side,
            )
            center_alignment = Alignment(horizontal="center", vertical="center")
            right_alignment = Alignment(horizontal="right", vertical="center")

            # 1. Aplicar bordes y centrado a todas las celdas
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = cell_border
                    cell.alignment = center_alignment

            # 2. Formato en negrita para la fila de cabeceras
            for cell in worksheet[1]:
                cell.font = header_font

            # 3. Ajuste automático del ancho de las columnas
            for col_idx, column_cells in enumerate(worksheet.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                for cell in column_cells:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                # Se le añade +2 de padding al texto más largo
                adjusted_width = max_length + 2
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # 4. Formato de dos decimales para la columna 'I' (IMPORTE NETO PAGAR)
            for cell in worksheet["I"][1:]:
                cell.number_format = "#,##0.00"
                cell.alignment = right_alignment

        # Extraemos los bytes
        excel_bytes = output_buffer.getvalue()

        # Generamos el nombre dinámico con la fecha de hoy
        filename = f"CapitalExpress_{datetime.now().strftime('%d%m%Y')}.xlsx"

        return filename, excel_bytes
